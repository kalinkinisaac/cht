from __future__ import annotations

import io
import logging
from typing import List

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from cht.table import Table

from .cluster_store import ClusterStore

logger = logging.getLogger("cht.api.services")


class MetadataService:
    """
    Service contract for ClickHouse metadata operations.

    Concrete implementations should delegate to Cluster/Table helpers rather than running ad-hoc
    SQL in the web layer.
    """

    def list_databases(
        self, *, cluster: str | None = None
    ) -> List[str]:  # pragma: no cover - interface
        raise NotImplementedError

    def list_tables(
        self, database: str, *, cluster: str | None = None
    ):  # pragma: no cover - interface
        raise NotImplementedError

    def list_columns(
        self, database: str, table: str, *, cluster: str | None = None
    ):  # pragma: no cover - interface
        raise NotImplementedError

    def update_table_comment(
        self, database: str, table: str, comment: str, *, cluster: str | None = None
    ) -> None:  # pragma: no cover - interface
        raise NotImplementedError

    def update_column_comment(
        self,
        database: str,
        table: str,
        column: str,
        comment: str,
        *,
        cluster: str | None = None,
    ) -> None:  # pragma: no cover - interface
        raise NotImplementedError

    def export_table_descriptions_to_excel(
        self, databases: list[str], *, cluster: str | None = None
    ) -> bytes:  # pragma: no cover - interface
        raise NotImplementedError


class ClickHouseMetadataService(MetadataService):
    """Concrete service that uses Cluster/Table for metadata operations."""

    def __init__(self, cluster_store: ClusterStore) -> None:
        self.cluster_store = cluster_store

    @staticmethod
    def _escape(value: str) -> str:
        return value.replace("'", "''")

    def _get_cluster(self, cluster: str | None):
        logger.info(f"Getting cluster: {cluster}")
        try:
            result = self.cluster_store.get_cluster(cluster)
            logger.info(f"Successfully got cluster: {result}")
            return result
        except Exception as e:
            logger.error(f"Error getting cluster {cluster}: {e}")
            raise

    def list_databases(self, *, cluster: str | None = None) -> List[str]:
        try:
            logger.info(f"Listing databases for cluster: {cluster}")
            cluster_obj = self._get_cluster(cluster)
            logger.info(f"Got cluster object: {cluster_obj}")
            rows = cluster_obj.query_with_fresh_client("SHOW DATABASES")
            logger.info(f"Query returned {len(rows or [])} databases")
            return [row[0] for row in rows or []]
        except Exception as e:
            logger.error(f"Error listing databases for cluster {cluster}: {e}")
            raise

    def list_tables(self, database: str, *, cluster: str | None = None):
        try:
            logger.info(f"Listing tables for database {database}, cluster: {cluster}")
            db = self._escape(database)
            cluster_obj = self._get_cluster(cluster)
            rows = cluster_obj.query_with_fresh_client(
                f"""
                SELECT name, comment, engine
                FROM system.tables
                WHERE database = '{db}'
                  AND (
                    engine LIKE '%MergeTree%' 
                    OR engine = 'MergeTree'
                    OR engine LIKE 'Replacing%MergeTree'
                    OR engine LIKE 'Summing%MergeTree'
                    OR engine LIKE 'Aggregating%MergeTree'
                    OR engine LIKE 'Collapsing%MergeTree'
                    OR engine LIKE 'VersionedCollapsing%MergeTree'
                    OR engine LIKE 'GraphiteMergeTree'
                  )
                ORDER BY name
                """
            )
            logger.info(f"Query returned {len(rows or [])} MergeTree tables")
            return [{"name": row[0], "comment": row[1] or None} for row in rows or []]
        except Exception as e:
            logger.error(f"Error listing tables for database {database}, cluster {cluster}: {e}")
            raise

    def list_columns(self, database: str, table: str, *, cluster: str | None = None):
        try:
            logger.info(f"Listing columns for {database}.{table}, cluster: {cluster}")
            db = self._escape(database)
            tbl = self._escape(table)
            cluster_obj = self._get_cluster(cluster)
            rows = cluster_obj.query_with_fresh_client(
                f"""
                SELECT name, type, comment
                FROM system.columns
                WHERE database = '{db}' AND table = '{tbl}'
                ORDER BY position
                """
            )
            logger.info(f"Query returned {len(rows or [])} columns")
            return [
                {"name": row[0], "type": row[1], "comment": row[2] or None} for row in rows or []
            ]
        except Exception as e:
            logger.error(f"Error listing columns for {database}.{table}, cluster {cluster}: {e}")
            raise

    def update_table_comment(
        self, database: str, table: str, comment: str, *, cluster: str | None = None
    ) -> None:
        table_obj = Table(database, table, cluster=self._get_cluster(cluster))
        table_obj.set_comment(comment)

    def update_column_comment(
        self,
        database: str,
        table: str,
        column: str,
        comment: str,
        *,
        cluster: str | None = None,
    ) -> None:
        table_obj = Table(database, table, cluster=self._get_cluster(cluster))
        table_obj.set_column_comment(column, comment)

    def export_table_descriptions_to_excel(
        self, databases: list[str], *, cluster: str | None = None
    ) -> bytes:
        """Export table descriptions for selected databases to Excel format.

        Creates one worksheet per table with columns: Column Name, Column Type, Comment.
        Each worksheet is named with the full table name (database.table).
        """
        try:
            logger.info(
                f"Exporting table descriptions for databases {databases}, cluster: {cluster}"
            )

            # Create a new workbook
            workbook = openpyxl.Workbook()
            # Remove the default worksheet
            workbook.remove(workbook.active)

            # Header style
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            # Process each database
            for database in databases:
                try:
                    # Get all tables for this database
                    tables = self.list_tables(database, cluster=cluster)
                    logger.info(f"Found {len(tables)} tables in database {database}")

                    for table_info in tables:
                        table_name = table_info["name"]
                        table_comment = table_info.get("comment", "")
                        full_table_name = f"{database}.{table_name}"

                        try:
                            # Get column information for this table
                            columns = self.list_columns(database, table_name, cluster=cluster)
                            logger.info(f"Found {len(columns)} columns in table {full_table_name}")

                            # Create worksheet for this table
                            # Truncate worksheet name if too long (Excel limit is 31 characters)
                            ws_name = full_table_name
                            if len(ws_name) > 31:
                                ws_name = ws_name[:28] + "..."

                            worksheet = workbook.create_sheet(title=ws_name)

                            # Add table info header
                            worksheet.merge_cells("A1:C1")
                            worksheet["A1"] = f"Table: {full_table_name}"
                            worksheet["A1"].font = Font(bold=True, size=14)
                            worksheet["A1"].alignment = Alignment(horizontal="center")

                            # Add description section with separate columns
                            worksheet["A2"] = "Description"
                            worksheet["A2"].font = Font(bold=True)
                            worksheet.merge_cells("B2:C2")
                            if table_comment:
                                worksheet["B2"] = table_comment
                            else:
                                worksheet["B2"] = "[Add table description here]"
                                worksheet["B2"].font = Font(color="888888")
                            worksheet["B2"].alignment = Alignment(horizontal="left", wrap_text=True)

                            # Add some spacing
                            worksheet.row_dimensions[3].height = 8  # Empty row for spacing
                            start_row = 4

                            # Add column headers with better styling
                            headers = ["Column Name", "Column Type", "Comment"]
                            for col, header in enumerate(headers, 1):
                                cell = worksheet.cell(row=start_row, column=col, value=header)
                                cell.font = header_font
                                cell.fill = header_fill
                                cell.alignment = header_alignment
                                cell.border = Border(
                                    left=Side(style="thin"),
                                    right=Side(style="thin"),
                                    top=Side(style="thin"),
                                    bottom=Side(style="thin"),
                                )

                            # Add column data with minimal formatting
                            for row, column in enumerate(columns, start_row + 1):
                                # Column name
                                worksheet.cell(row=row, column=1, value=column["name"])

                                # Column type
                                worksheet.cell(row=row, column=2, value=column["type"])

                                # Column comment (with placeholder if empty)
                                comment_value = column.get("comment")
                                comment_cell = worksheet.cell(row=row, column=3)
                                if comment_value:
                                    comment_cell.value = comment_value
                                else:
                                    comment_cell.value = "[Add column comment here]"
                                    comment_cell.font = Font(color="AAAAAA")

                                # Add minimal borders to all data cells
                                for col in range(1, 4):
                                    cell = worksheet.cell(row=row, column=col)
                                    cell.border = Border(
                                        left=Side(style="thin"),
                                        right=Side(style="thin"),
                                        top=Side(style="thin"),
                                        bottom=Side(style="thin"),
                                    )
                                    cell.alignment = Alignment(vertical="top", wrap_text=True)

                            # Auto-adjust column widths
                            for col_num, column_cells in enumerate(worksheet.columns, 1):
                                max_length = 0
                                column_letter = openpyxl.utils.get_column_letter(col_num)

                                for cell in column_cells:
                                    try:
                                        # Skip merged cells which don't have individual values
                                        if hasattr(cell, "value") and cell.value is not None:
                                            cell_length = len(str(cell.value))
                                            if cell_length > max_length:
                                                max_length = cell_length
                                    except Exception:
                                        # Skip any problematic cells
                                        pass

                                # Set column width with reasonable limits
                                adjusted_width = min(
                                    max(max_length + 2, 10), 50
                                )  # Min 10, Max 50 characters
                                worksheet.column_dimensions[column_letter].width = adjusted_width

                        except Exception as e:
                            logger.error(f"Error processing table {full_table_name}: {e}")
                            # Continue with other tables
                            continue

                except Exception as e:
                    logger.error(f"Error processing database {database}: {e}")
                    # Continue with other databases
                    continue

            # If no worksheets were created, create a summary sheet
            if not workbook.worksheets:
                worksheet = workbook.create_sheet(title="Export Summary")
                worksheet["A1"] = "No tables found in the selected databases"
                worksheet["A1"].font = Font(bold=True)

            # Save to bytes
            excel_buffer = io.BytesIO()
            workbook.save(excel_buffer)
            excel_buffer.seek(0)

            logger.info(
                f"Successfully created Excel export with {len(workbook.worksheets)} worksheets"
            )
            return excel_buffer.getvalue()

        except Exception as e:
            logger.error(f"Error creating Excel export: {e}")
            raise
